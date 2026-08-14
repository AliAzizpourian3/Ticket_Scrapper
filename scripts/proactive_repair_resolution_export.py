#!/usr/bin/env python3
"""Export proactive ISP repair evidence from Tickets_Trimmed JSON files to Excel.

Run from the repository root:
    pip install openpyxl
    python scripts/proactive_repair_resolution_export.py --include-candidates

By default the script reads:
    Database/Ticket_Extraction/Tickets_Trimmed
and creates:
    Database/Ticket_Extraction/proactive_repair_resolution_summary.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Missing dependency. Install it with: pip install openpyxl") from exc

PROACTIVE = re.compile(r"\bproactive\b|proactive monitoring activity", re.I)
REPAIR_TYPES = {
    "Bender / IMD": [r"\bbender\b", r"\bimd\b", r"insulation monitoring"],
    "Contactor": [r"\bcontactor(?:s)?\b", r"\bqa\d{1,2}[-/ ]?\d{1,2}\b"],
    "Cable / Connector": [r"cabletempsensor", r"\bccs cable\b", r"\bcable replacement\b", r"\bconnector replacement\b", r"\bplug replacement\b"],
    "Fan": [r"\bmain fan\b", r"\bcooling fan\b", r"\bfan replacement\b"],
    "Power converter": [r"\bpower converter\b", r"\bconverter replacement\b"],
    "Voltage transducer": [r"\bvoltage transducer\b", r"\btransducer replacement\b"],
    "Other component": [r"component exchange", r"spare part"],
}
COMPLETED = [
    r"\b(?:was|were|has been|have been|is|are) replaced\b",
    r"\breplaced (?:the )?(?:main )?(?:fan|bender|imd|contactor|cable|connector|plug|converter|component|part|module|unit)\b",
    r"\b(?:repair|replacement|component exchange) (?:was )?(?:completed|performed|successful|successfully completed)\b",
    r"\b(?:onsite|on-site) repair (?:was )?(?:completed|performed|successful)\b",
    r"\bissue (?:has been |was )?(?:resolved|fixed)\b",
    r"\bcharger (?:is|was) working (?:again|now)\b",
]
RECOMMENDED = [
    r"\bplease proceed (?:with|to) (?:the )?replacement\b",
    r"\b(?:need|needs|recommend|suggest) (?:to )?(?:replace|replacement)\b",
    r"\bconsider replacing\b",
    r"\bspare part ordering\b",
    r"\bon[- ]site repair.*\byes\b",
    r"\bcomponent exchange.*\byes\b",
]
NOT_COMPLETED = [
    r"\bno need (?:to |for )?(?:replace|replacement)\b",
    r"\bdon['’]t need (?:to |for )?(?:replace|replacement)\b",
    r"\bwill (?:be )?replace(?:d)?\b",
    r"\bto be replaced\b",
    r"\bawaiting (?:replacement|repair|parts)\b",
    r"\bplanned (?:replacement|repair)\b",
]
HEADERS = [
    "Ticket Number", "Title", "Created", "Last Activity", "Stage", "Priority", "Charger",
    "Serial / Main Asset", "Case Nature", "Description", "Summary", "Diagnostics", "Resolution",
    "Repair Type", "Repair Completed", "Evidence", "Evidence Date", "JSON File",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def clean(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\u00a0", " ")).strip()


def any_match(patterns: list[str], value: str) -> re.Match[str] | None:
    for pattern in patterns:
        match = re.search(pattern, value, re.I | re.S)
        if match:
            return match
    return None


def excerpt(value: str, match: re.Match[str] | None, radius: int = 350) -> str:
    if not match:
        return ""
    return clean(value[max(0, match.start() - radius): min(len(value), match.end() + radius)])


def shorten(value: str, size: int = 600) -> str:
    value = clean(value)
    return value if len(value) <= size else value[:size - 3].rstrip() + "..."


def field(ticket: dict[str, Any], *keys: str) -> str:
    for key in keys:
        if ticket.get(key) not in (None, ""):
            return text(ticket[key])
    return ""


def entries(ticket: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ("conversation", "conversations", "activities", "timeline", "emails", "messages", "notes"):
        if isinstance(ticket.get(key), list):
            return [item for item in ticket[key] if isinstance(item, dict)]
    return []


def entry_body(entry: dict[str, Any]) -> str:
    return clean("\n".join(text(entry.get(key)) for key in ("subject", "body", "description", "text", "note", "content") if entry.get(key)))


def entry_date(entry: dict[str, Any]) -> str:
    return field(entry, "date", "created_on", "createdAt", "timestamp", "modifiedon")


def repair_type(value: str) -> str:
    result = [name for name, patterns in REPAIR_TYPES.items() if any_match(patterns, value)]
    return "; ".join(result) if result else "Unspecified repair"


def record(ticket: dict[str, Any], source: Path) -> dict[str, str] | None:
    title = field(ticket, "title", "subject", "name")
    description = field(ticket, "description", "case_description", "details")
    conversation = [(entry_body(item), entry_date(item)) for item in entries(ticket)]
    combined = clean("\n".join([title, description, *(body for body, _ in conversation)]))

    if not PROACTIVE.search(combined):
        return None

    completed_evidence: list[tuple[str, str]] = []
    for body, date in conversation:
        hit = any_match(COMPLETED, body)
        if hit and not any_match(NOT_COMPLETED, excerpt(body, hit, 180)):
            completed_evidence.append((excerpt(body, hit), date))

    top_hit = any_match(COMPLETED, description)
    if top_hit and not any_match(NOT_COMPLETED, excerpt(description, top_hit, 180)):
        completed_evidence.append((excerpt(description, top_hit), field(ticket, "last_activity", "modified_on")))

    recommendation = any_match(RECOMMENDED, combined)
    related_repair = completed_evidence or recommendation or any(any_match(patterns, combined) for patterns in REPAIR_TYPES.values())
    if not related_repair:
        return None

    if completed_evidence:
        evidence, evidence_date = completed_evidence[-1]
        repair_completed = "Yes"
    else:
        evidence = excerpt(combined, recommendation) if recommendation else "Repair-related wording found but no explicit completion evidence."
        evidence_date = ""
        repair_completed = "No - candidate only"

    diagnostics = ""
    diagnosis = re.search(r"(?:diagnostics?|troubleshooting)\s*[:\-]?(.*?)(?:(?:resolution|solution|action)\s*[:\-]|$)", combined, re.I | re.S)
    if diagnosis:
        diagnostics = shorten(diagnosis.group(1))

    return {
        "Ticket Number": field(ticket, "ticket_number", "case_number", "number", "incident_id"),
        "Title": title,
        "Created": field(ticket, "created_on", "createdAt", "created", "createdon"),
        "Last Activity": field(ticket, "last_activity", "lastActivity", "modified_on", "modifiedon"),
        "Stage": field(ticket, "stage", "status", "state"),
        "Priority": field(ticket, "priority"),
        "Charger": field(ticket, "chargerID", "charger_id", "charger"),
        "Serial / Main Asset": field(ticket, "serial", "serial_number", "main_asset", "asset"),
        "Case Nature": "Proactive",
        "Description": description,
        "Summary": shorten(description or title),
        "Diagnostics": diagnostics,
        "Resolution": shorten(evidence),
        "Repair Type": repair_type(combined),
        "Repair Completed": repair_completed,
        "Evidence": evidence,
        "Evidence Date": evidence_date,
        "JSON File": source.name,
    }


def collect(input_dir: Path) -> tuple[list[dict[str, str]], list[str]]:
    rows, errors = [], []
    for path in sorted(input_dir.glob("*.json")):
        try:
            with path.open("r", encoding="utf-8-sig") as handle:
                ticket = json.load(handle)
            if not isinstance(ticket, dict):
                errors.append(f"{path.name}: root JSON is not an object")
                continue
            result = record(ticket, path)
            if result:
                rows.append(result)
        except (OSError, UnicodeDecodeError, json.JSONDecodeError) as error:
            errors.append(f"{path.name}: {error}")
    return rows, errors


def add_data_sheet(book: Workbook, name: str, rows: list[dict[str, str]]) -> None:
    sheet = book.create_sheet(name)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32
    widths = [15, 46, 22, 22, 18, 10, 18, 28, 15, 55, 55, 55, 55, 24, 21, 75, 22, 42]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    col = get_column_letter(HEADERS.index("Repair Completed") + 1)
    sheet.conditional_formatting.add(f"{col}2:{col}{max(2, sheet.max_row)}", CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="C6EFCE")))


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description="Export proactive ISP repair evidence from Tickets_Trimmed JSON files.")
    parser.add_argument("--input-dir", type=Path, default=root / "Database" / "Ticket_Extraction" / "Tickets_Trimmed")
    parser.add_argument("--output", type=Path, default=root / "Database" / "Ticket_Extraction" / "proactive_repair_resolution_summary.xlsx")
    parser.add_argument("--include-candidates", action="store_true")
    args = parser.parse_args()
    if not args.input_dir.is_dir():
        raise SystemExit(f"Input directory does not exist: {args.input_dir}")

    rows, errors = collect(args.input_dir)
    confirmed = sorted((row for row in rows if row["Repair Completed"] == "Yes"), key=lambda row: (row["Created"], row["Ticket Number"]), reverse=True)
    candidates = sorted((row for row in rows if row["Repair Completed"] != "Yes"), key=lambda row: (row["Created"], row["Ticket Number"]), reverse=True)

    book = Workbook()
    overview = book.active
    overview.title = "Overview"
    overview.append(["Metric", "Value"])
    overview.append(["Confirmed proactive repairs", len(confirmed)])
    overview.append(["Proactive repair candidates awaiting confirmation", len(candidates)])
    overview.append(["Unreadable / invalid JSON files", len(errors)])
    overview.append([])
    overview.append(["Confirmed repair type", "Count"])
    for name, count in sorted(Counter(row["Repair Type"] for row in confirmed).items()):
        overview.append([name, count])
    if errors:
        overview.append([])
        overview.append(["Load errors", "Details"])
        for error in errors:
            overview.append(["JSON error", error])
    for cell in overview[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    overview.column_dimensions["A"].width = 48
    overview.column_dimensions["B"].width = 110
    for row in overview.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    add_data_sheet(book, "Confirmed Repairs", confirmed)
    if args.include_candidates:
        add_data_sheet(book, "Candidates - Review", candidates)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    book.save(args.output)
    print(f"Created: {args.output}")
    print(f"Confirmed proactive repairs: {len(confirmed)}")
    print(f"Candidates requiring review: {len(candidates)}")
    if errors:
        print(f"JSON files skipped due to errors: {len(errors)}")


if __name__ == "__main__":
    main()
